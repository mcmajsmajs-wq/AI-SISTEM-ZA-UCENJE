# -*- coding: utf-8 -*-
"""
================================================================================
XLSX Export Service
================================================================================
Kreira Excel tabelu od prevedenih chunk-ova.

Verzija: 1.0.0
================================================================================
"""

import io
import logging
from typing import List, Dict, Any
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from app.services.base_export_service import BaseExportService

logger = logging.getLogger(__name__)


class XLSXExportService(BaseExportService):
    """Service za export dokumenta u XLSX format."""

    def generate(
        self,
        title: str,
        chunks: List[Dict[str, Any]],
        include_original: bool = False,
    ) -> bytes:
        """
        Generiše XLSX dokument od chunk-ova.

        Args:
            title: Naslov dokumenta
            chunks: Lista chunk-ova (dictionaries)
            include_original: Da li uključuje originalni tekst

        Returns:
            XLSX bytes
        """
        self._report_progress(0, 100, "Priprema XLSX generisanja...")
        wb = Workbook()

        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        # Main content sheet
        ws = wb.create_sheet("Sadržaj", 0)

        # Header styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Title row
        ws.merge_cells("A1:E1")
        title_cell = ws["A1"]
        title_cell.value = title
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Column headers
        headers = ["#", "Naslov/Podnaslov", "Sadržaj", "Stranica", "Original"]
        if not include_original:
            headers = headers[:4]

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        self._report_progress(10, 100, "Dodavanje podataka...")

        # Data rows
        row_num = 4
        for idx, chunk in enumerate(chunks, 1):
            # Sequence number
            ws.cell(row=row_num, column=1, value=idx).border = thin_border

            # Heading/Title
            heading = chunk.get("parent_heading") or chunk.get("heading_level", "")
            ws.cell(row=row_num, column=2, value=str(heading)).border = thin_border

            content = self._get_content(chunk)

            # Truncate if too long
            if len(content) > 500:
                content = content[:500] + "..."

            ws.cell(row=row_num, column=3, value=content).border = thin_border

            # Page number
            page = chunk.get("page_number", "")
            ws.cell(
                row=row_num, column=4, value=str(page) if page else ""
            ).border = thin_border

            # Original content
            if include_original:
                original = chunk.get("content", "")
                if len(original) > 300:
                    original = original[:300] + "..."
                ws.cell(row=row_num, column=5, value=original).border = thin_border

            row_num += 1

            # Progress update
            if idx % 500 == 0:
                progress = 10 + int((idx / len(chunks)) * 80)
                self._report_progress(progress, 100, f"Podaci: {idx}/{len(chunks)}...")

        # Adjust column widths
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 60
        ws.column_dimensions["D"].width = 10
        if include_original:
            ws.column_dimensions["E"].width = 40

        # Summary sheet
        ws_summary = wb.create_sheet("Pregled", 0)
        ws_summary["A1"] = "Pregled dokumenta"
        ws_summary["A1"].font = Font(bold=True, size=14)

        ws_summary["A3"] = "Naslov:"
        ws_summary["B3"] = title

        ws_summary["A4"] = "Broj sekcija:"
        ws_summary["B4"] = len(chunks)

        ws_summary["A5"] = "Datum generisanja:"
        ws_summary["B5"] = datetime.now().strftime("%Y-%m-%d %H:%M")

        # Adjust column width
        ws_summary.column_dimensions["A"].width = 20
        ws_summary.column_dimensions["B"].width = 40

        self._report_progress(95, 100, "Finalizacija XLSX-a...")

        # Save to bytes
        xlsx_bytes = io.BytesIO()
        wb.save(xlsx_bytes)
        xlsx_bytes.seek(0)

        self._report_progress(100, 100, "XLSX generisanje završeno!")
        return xlsx_bytes.getvalue()


xlsx_export_service = XLSXExportService()
